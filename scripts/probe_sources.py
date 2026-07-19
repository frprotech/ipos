#!/usr/bin/env python3
"""Diagnostic: probe candidate CSE data sources from a GitHub runner.
Temporary tool — not part of the site."""

import io
import json
import re
import requests

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
    ),
    "Accept": "application/json, text/html;q=0.9, */*;q=0.8",
}

# 1) Inspect the "Stock List Changes" workbook — it usually carries new
#    listings/delistings with effective dates.
print("=" * 100)
url = "https://market-reports.thecse.com/Deprecated/CSE%20Stock%20List%20Changes.xlsx"
print("URL:", url)
try:
    from openpyxl import load_workbook
    r = requests.get(url, headers=HEADERS, timeout=30)
    print("STATUS:", r.status_code, "| LEN:", len(r.content))
    wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    for ws in wb.worksheets:
        print("SHEET:", ws.title)
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            print("  ROW", i, ":", [str(c)[:40] if c is not None else "" for c in row[:12]])
            if i >= 14:
                break
except Exception as exc:
    print("ERROR:", exc)

# 2) The thecse.com Next.js data route for the listed-companies page.
print("=" * 100)
page_url = "https://thecse.com/listing/listed-companies/"
print("URL:", page_url, "(discover buildId, then fetch _next data)")
try:
    html = requests.get(page_url, headers=HEADERS, timeout=30).text
    m = re.search(r'"buildId"\s*:\s*"([^"]+)"', html)
    print("BUILD-ID:", m.group(1) if m else None)
    if m:
        data_url = f"https://thecse.com/_next/data/{m.group(1)}/listing/listed-companies.json"
        r = requests.get(data_url, headers=HEADERS, timeout=30)
        print("DATA-URL:", data_url)
        print("STATUS:", r.status_code, "| TYPE:", r.headers.get("content-type"),
              "| LEN:", len(r.content))
        if r.status_code == 200:
            payload = r.json()

            def walk(obj, path="", depth=0):
                """Print the shape of the JSON: keys and first list items."""
                if depth > 6:
                    return
                if isinstance(obj, dict):
                    for k, v in list(obj.items())[:25]:
                        if isinstance(v, (dict, list)):
                            print("  " * depth + f"{path}.{k} ({type(v).__name__}, "
                                  f"len={len(v)})")
                            walk(v, f"{path}.{k}", depth + 1)
                elif isinstance(obj, list) and obj:
                    first = obj[0]
                    if isinstance(first, dict):
                        print("  " * depth + f"{path}[0] keys: {list(first.keys())[:20]}")
                        print("  " * depth + f"{path}[0] sample: "
                              + json.dumps(first, default=str)[:600])

            walk(payload)
except Exception as exc:
    print("ERROR:", exc)

# 3) Direct guesses at an API the CSE site might expose.
for url in [
    "https://thecse.com/api/listings",
    "https://thecse.com/api/companies",
]:
    print("=" * 100)
    print("URL:", url)
    try:
        r = requests.get(url, headers=HEADERS, timeout=30)
        print("STATUS:", r.status_code, "| TYPE:", r.headers.get("content-type"),
              "| LEN:", len(r.content))
        if r.status_code == 200:
            print(r.text[:1500])
    except Exception as exc:
        print("ERROR:", exc)
