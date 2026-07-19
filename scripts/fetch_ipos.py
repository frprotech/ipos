#!/usr/bin/env python3
"""Fetch IPO / new-listing data from public exchange sources and write data/ipos.json.

Covered exchanges:
  - NASDAQ + NYSE (+ NYSE American)  via the Nasdaq IPO calendar API
  - TSX / TSX Venture                via tsx.com "Current listing activity" page
  - ASX                              via asx.com.au upcoming floats & listings
  - CSE                              via the CSE public securities feed

Each exchange fetcher is isolated: if one source fails or changes its markup,
the previous data for that exchange is kept and the rest still update.
"""

from __future__ import annotations

import datetime as dt
import json
import re
import sys
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from dateutil import parser as dateparser

ROOT = Path(__file__).resolve().parent.parent
DATA_FILE = ROOT / "data" / "ipos.json"

TIMEOUT = 30
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
    ),
    "Accept": "application/json, text/html;q=0.9, */*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

TODAY = dt.date.today()
# Keep listings from the past year plus anything upcoming.
KEEP_AFTER = TODAY - dt.timedelta(days=365)


def http_get(url: str, **kwargs) -> requests.Response:
    resp = requests.get(url, headers=HEADERS, timeout=TIMEOUT, **kwargs)
    resp.raise_for_status()
    return resp


def parse_date(value) -> str | None:
    """Best-effort parse of a date-ish value to ISO YYYY-MM-DD."""
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"tba", "tbd", "n/a", "-", "--"}:
        return None
    # Strip common noise: "(expected)" notes, footnote markers ("#"),
    # times ("12.00pm") and timezone tags, as seen on the ASX page.
    text = re.sub(r"\(.*?\)", " ", text)
    text = re.sub(r"#+", " ", text)
    text = re.sub(r"\b\d{1,2}[.:]\d{2}\s*(?:am|pm)?\b", " ", text, flags=re.I)
    text = re.sub(r"\b(?:AEST|AEDT|EST|EDT|ET|noon|midday)\b", " ", text, flags=re.I)
    text = re.sub(r"\s+", " ", text).strip()
    try:
        return dateparser.parse(text, dayfirst=False, fuzzy=True).date().isoformat()
    except (ValueError, OverflowError):
        try:
            return dateparser.parse(text, dayfirst=True, fuzzy=True).date().isoformat()
        except (ValueError, OverflowError):
            return None


def record(exchange: str, company: str, ticker: str = "", listing_date: str | None = None,
           sector: str = "", status: str = "", source: str = "") -> dict | None:
    company = re.sub(r"\s+", " ", str(company or "")).strip()
    ticker = str(ticker or "").strip().upper()
    if not company:
        return None
    return {
        "company": company,
        "ticker": ticker,
        "exchange": exchange,
        "listing_date": listing_date,
        "sector": re.sub(r"\s+", " ", str(sector or "")).strip(),
        "status": status,
        "source": source,
        "fetched_at": TODAY.isoformat(),
    }


# ---------------------------------------------------------------------------
# NASDAQ + NYSE — Nasdaq's IPO calendar API covers all US listings and tags
# each deal with its proposed exchange.
# ---------------------------------------------------------------------------

def fetch_us() -> list[dict]:
    source = "https://www.nasdaq.com/market-activity/ipos"
    months = []
    cursor = TODAY.replace(day=1) - dt.timedelta(days=1)  # previous month
    cursor = cursor.replace(day=1)
    for _ in range(4):  # previous, current and next two months
        months.append(cursor.strftime("%Y-%m"))
        cursor = (cursor + dt.timedelta(days=32)).replace(day=1)

    def map_exchange(raw: str) -> str | None:
        raw = (raw or "").upper()
        if "NASDAQ" in raw:
            return "NASDAQ"
        if "AMERICAN" in raw or "AMEX" in raw:
            return "NYSE American"
        if "NYSE" in raw:
            return "NYSE"
        return None

    out: list[dict] = []
    for month in months:
        data = http_get(
            f"https://api.nasdaq.com/api/ipo/calendar?date={month}"
        ).json().get("data") or {}

        sections = [
            ("priced", (data.get("priced") or {}).get("rows"), "pricedDate", "Priced"),
            ("upcoming", ((data.get("upcoming") or {}).get("upcomingTable") or {}).get("rows"),
             "expectedPriceDate", "Expected"),
            ("filed", (data.get("filed") or {}).get("rows"), "filedDate", "Filed"),
        ]
        for _name, rows, date_key, status in sections:
            for row in rows or []:
                exchange = map_exchange(row.get("proposedExchange") or row.get("exchange"))
                if not exchange:
                    continue
                rec = record(
                    exchange=exchange,
                    company=row.get("companyName"),
                    ticker=row.get("proposedTickerSymbol") or row.get("symbol"),
                    listing_date=parse_date(row.get(date_key)),
                    status=status,
                    source=source,
                )
                if rec:
                    out.append(rec)
    if not out:
        raise RuntimeError("Nasdaq IPO calendar returned no usable rows")
    return out


# ---------------------------------------------------------------------------
# Generic HTML table extraction — used for exchange pages that publish their
# listing activity as plain HTML tables (TSX, ASX fallback).
# ---------------------------------------------------------------------------

COMPANY_KEYS = ("company", "issuer", "name", "entity")
TICKER_KEYS = ("symbol", "ticker", "code", "stock")
DATE_KEYS = ("date", "listed", "listing", "expected", "anticipated")
SECTOR_KEYS = ("sector", "industry", "category", "type of listing", "activity")


def _match_col(headers: list[str], keys: tuple[str, ...]) -> int | None:
    for i, h in enumerate(headers):
        if any(k in h for k in keys):
            return i
    return None


def extract_table_rows(html: str, exchange: str, source: str, status: str) -> list[dict]:
    soup = BeautifulSoup(html, "html.parser")
    out: list[dict] = []
    for table in soup.find_all("table"):
        header_row = table.find("tr")
        if not header_row:
            continue
        headers = [c.get_text(" ", strip=True).lower()
                   for c in header_row.find_all(["th", "td"])]
        ci_company = _match_col(headers, COMPANY_KEYS)
        ci_date = _match_col(headers, DATE_KEYS)
        if ci_company is None or ci_date is None:
            continue
        ci_ticker = _match_col(headers, TICKER_KEYS)
        ci_sector = _match_col(headers, SECTOR_KEYS)

        for tr in header_row.find_next_siblings("tr"):
            cells = [c.get_text(" ", strip=True) for c in tr.find_all(["td", "th"])]
            if len(cells) <= max(ci_company, ci_date):
                continue
            rec = record(
                exchange=exchange,
                company=cells[ci_company],
                ticker=cells[ci_ticker] if ci_ticker is not None and ci_ticker < len(cells) else "",
                listing_date=parse_date(cells[ci_date]),
                sector=cells[ci_sector] if ci_sector is not None and ci_sector < len(cells) else "",
                status=status,
                source=source,
            )
            if rec:
                out.append(rec)
    return out


# ---------------------------------------------------------------------------
# TSX / TSX Venture — tsx.com publishes new company listings (TSX and TSXV
# combined) as a server-rendered Date | Company table, where the company cell
# reads "Company Name (TICKER)".
# ---------------------------------------------------------------------------

def fetch_tsx() -> list[dict]:
    url = "https://www.tsx.com/en/news/new-company-listings"
    soup = BeautifulSoup(http_get(url).text, "html.parser")
    out: list[dict] = []
    for table in soup.find_all("table"):
        headers = [c.get_text(" ", strip=True).lower()
                   for c in table.find_all(["th"])]
        if "date" not in headers or "company" not in headers:
            continue
        for tr in table.find_all("tr"):
            cells = tr.find_all("td")
            if len(cells) < 2:
                continue
            date = parse_date(cells[0].get_text(" ", strip=True))
            name = cells[1].get_text(" ", strip=True).replace("\xa0", " ")
            m = re.match(r"^(.*?)\s*\(([^()]{1,12})\)\s*$", name)
            company, ticker = (m.group(1), m.group(2)) if m else (name, "")
            rec = record(
                exchange="TSX",
                company=company,
                ticker=ticker,
                listing_date=date,
                status="Listed",
                source=url,
            )
            if rec:
                out.append(rec)
    if not out:
        raise RuntimeError("No new-company-listings table found on tsx.com")
    return out


# ---------------------------------------------------------------------------
# ASX — the upcoming floats page renders one key/value detail table per
# company (Listing date / Security code / Principal activities / ...), with
# the company name in the nearest preceding heading.
# ---------------------------------------------------------------------------

ASX_PAGE = "https://www.asx.com.au/listings/upcoming-floats-and-listings"

ASX_FIELD_LABELS = {"listing date", "contact details", "principal activities",
                    "issue price", "issue type", "security code",
                    "capital to be raised", "expected offer close date",
                    "underwriter"}


def fetch_asx() -> list[dict]:
    soup = BeautifulSoup(http_get(ASX_PAGE).text, "html.parser")
    out: list[dict] = []
    for table in soup.find_all("table"):
        kv: dict[str, str] = {}
        for tr in table.find_all("tr"):
            cells = tr.find_all("td")
            if len(cells) >= 2:
                key = cells[0].get_text(" ", strip=True).lower().rstrip(":")
                kv[key] = cells[1].get_text(" ", strip=True)
        if "listing date" not in kv or "security code" not in kv:
            continue  # not a company detail table

        company = ""
        heading = table.find_previous(["h2", "h3", "h4", "h5"])
        if heading:
            text = heading.get_text(" ", strip=True)
            if text and text.lower() not in ASX_FIELD_LABELS and len(text) < 120:
                company = text
        rec = record(
            exchange="ASX",
            company=company or kv.get("security code", ""),
            ticker=kv.get("security code", ""),
            listing_date=parse_date(kv.get("listing date")),
            sector=kv.get("principal activities", ""),
            status="Upcoming",
            source=ASX_PAGE,
        )
        if rec:
            out.append(rec)
    if not out:
        raise RuntimeError("No company detail tables found on the ASX floats page")
    return out


# ---------------------------------------------------------------------------
# CSE — the CSE publishes its full stock list as an Excel file that includes
# each security's listing date; keep the ones listed in the past year.
# ---------------------------------------------------------------------------

def fetch_cse() -> list[dict]:
    import io

    from openpyxl import load_workbook

    url = "https://thecse.com/sites/default/files/CSE_Stock_List.xlsx"
    wb = load_workbook(io.BytesIO(http_get(url).content), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    header = None
    for row in rows:
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if any("symbol" in c for c in cells) and any("company" in c or "name" in c for c in cells):
            header = cells
            break
    if header is None:
        raise RuntimeError("Could not find header row in CSE stock list")

    def col(*keys):
        for i, h in enumerate(header):
            if any(k in h for k in keys):
                return i
        return None

    ci_symbol = col("symbol", "ticker")
    ci_company = col("company", "name", "security")
    ci_date = col("date listed", "listing date", "listed", "list date")
    ci_sector = col("industry", "sector")
    if ci_date is None:
        raise RuntimeError(f"CSE stock list has no listing-date column (header: {header})")

    out: list[dict] = []
    for row in rows:
        def cell(i):
            if i is None or i >= len(row) or row[i] is None:
                return ""
            v = row[i]
            return v.date().isoformat() if isinstance(v, dt.datetime) else str(v)

        listing_date = parse_date(cell(ci_date))
        if not listing_date or listing_date < KEEP_AFTER.isoformat():
            continue  # only recent listings belong in an IPO table
        rec = record(
            exchange="CSE",
            company=cell(ci_company),
            ticker=cell(ci_symbol),
            listing_date=listing_date,
            sector=cell(ci_sector),
            status="Listed",
            source="https://thecse.com/listing/listed-companies/",
        )
        if rec:
            out.append(rec)
    if not out:
        raise RuntimeError("CSE stock list yielded no recent listings")
    return out


# ---------------------------------------------------------------------------
# Merge / prune / write
# ---------------------------------------------------------------------------

FETCHERS: dict[str, tuple] = {
    # group name -> (fetcher, exchanges the group is authoritative for)
    "US": (fetch_us, {"NASDAQ", "NYSE", "NYSE American"}),
    "TSX": (fetch_tsx, {"TSX", "TSXV"}),
    "ASX": (fetch_asx, {"ASX"}),
    "CSE": (fetch_cse, {"CSE"}),
}


def keep(rec: dict) -> bool:
    date = rec.get("listing_date")
    if date:
        return date >= KEEP_AFTER.isoformat()
    # Undated rows (TBA listings) are kept while their source still reports
    # them; drop them once they go stale.
    fetched = rec.get("fetched_at") or "1970-01-01"
    return fetched >= (TODAY - dt.timedelta(days=60)).isoformat()


def main() -> int:
    existing: list[dict] = []
    if DATA_FILE.exists():
        try:
            existing = json.loads(DATA_FILE.read_text()).get("ipos") or []
        except json.JSONDecodeError:
            print("WARNING: existing data file is invalid JSON; starting fresh")

    merged: list[dict] = []
    failures: list[str] = []
    for name, (fetcher, exchanges) in FETCHERS.items():
        try:
            rows = fetcher()
            print(f"[{name}] fetched {len(rows)} rows")
            merged.extend(rows)
        except Exception as exc:  # keep last good data for this exchange group
            print(f"WARNING: [{name}] fetch failed ({exc}); keeping previous data")
            failures.append(name)
            merged.extend(r for r in existing if r.get("exchange") in exchanges)

    # Dedupe by exchange + ticker (or company when no ticker), newest wins.
    deduped: dict[tuple, dict] = {}
    for rec in merged:
        key = (rec.get("exchange"), rec.get("ticker") or rec.get("company", "").lower())
        prev = deduped.get(key)
        if prev is None or (rec.get("fetched_at") or "") >= (prev.get("fetched_at") or ""):
            deduped[key] = rec

    final = sorted(
        (r for r in deduped.values() if keep(r)),
        key=lambda r: (r.get("listing_date") or "9999-99-99"),
        reverse=True,
    )

    DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
    DATA_FILE.write_text(json.dumps({
        "updated_at": dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds"),
        "sources_failed": failures,
        "ipos": final,
    }, indent=2) + "\n")
    print(f"Wrote {len(final)} records to {DATA_FILE} "
          f"({len(failures)} source(s) failed: {failures or 'none'})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
